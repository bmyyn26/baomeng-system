"""
Excel 資料匯入腳本
一次性執行，將現有 Excel 資料匯入 Supabase 資料庫

執行方式：
  python import_data.py

需要先設定好 .env 或環境變數中的 SUPABASE_URL 和 SUPABASE_KEY
"""

import os
import sys
from datetime import datetime
from dotenv import load_dotenv
import openpyxl

load_dotenv()

from supabase import create_client, Client

SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ 請先設定 SUPABASE_URL 和 SUPABASE_KEY 環境變數")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Excel 檔案路徑（改成你實際的路徑）──────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MEMBERS_FILE    = os.path.join(BASE_DIR, '寶萌會員2_整理完成.xlsx')
BLACKLIST_FILE  = os.path.join(BASE_DIR, '黑名單3.xlsx')
ORDERS_FILE     = os.path.join(BASE_DIR, '訂單資料.xlsx')


def safe_str(val) -> str:
    """安全轉成字串，None 回傳空字串"""
    if val is None:
        return ''
    return str(val).strip()


def safe_int(val):
    """安全轉成整數，失敗回傳 None"""
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def safe_float(val):
    """安全轉成浮點數，失敗回傳 None"""
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


# ─────────────────────────────────────────────
#  匯入會員（寶萌會員2_整理完成.xlsx）
# ─────────────────────────────────────────────

def import_members():
    print("\n📂 匯入會員資料...")
    wb = openpyxl.load_workbook(MEMBERS_FILE, read_only=True, data_only=True)
    ws = wb.active

    batch = []
    seen_ids = set()  # 用 nickname_id 去重

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # 欄位：會員分類, 9188編號, 暱稱編號, 暱稱, 帳號, 戶名, 電話, 身分證, 備註, 備註2
        classification = safe_str(row[0])
        id_9188        = safe_int(row[1])
        nickname_id    = safe_int(row[2])
        nickname       = safe_str(row[3])
        bank_account   = safe_str(row[4])
        account_holder = safe_str(row[5])
        phone          = safe_str(row[6])
        id_number      = safe_str(row[7])
        notes          = safe_str(row[8])
        notes2         = safe_str(row[9])

        # 跳過空行（暱稱和 9188 編號都沒有）
        if not nickname and not id_9188:
            continue

        # 用 (id_9188, nickname_id) 組合去重
        key = (id_9188, nickname_id)
        if key in seen_ids:
            continue
        seen_ids.add(key)

        record = {
            'classification': classification or None,
            'id_9188':        id_9188,
            'nickname_id':    nickname_id,
            'nickname':       nickname or None,
            'bank_account':   bank_account or None,
            'account_holder': account_holder or None,
            'phone':          phone or None,
            'id_number':      id_number or None,
            'notes':          notes or None,
            'notes2':         notes2 or None,
            'is_blacklist':   False,
        }
        batch.append(record)

        # 每 100 筆送一次
        if len(batch) >= 100:
            supabase.table('customers').insert(batch).execute()
            print(f'  ✅ 已匯入 {len(batch)} 筆會員')
            batch = []

    if batch:
        supabase.table('customers').insert(batch).execute()
        print(f'  ✅ 已匯入 {len(batch)} 筆會員')

    wb.close()
    print(f'  🎉 會員匯入完成，共 {len(seen_ids)} 筆')


# ─────────────────────────────────────────────
#  匯入黑名單（黑名單3.xlsx）
# ─────────────────────────────────────────────

def import_blacklist():
    print("\n📂 匯入黑名單資料...")
    wb = openpyxl.load_workbook(BLACKLIST_FILE, read_only=True, data_only=True)
    ws = wb.active

    batch = []
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 欄位：會員分類, 會員編號, 暱稱編號, 暱稱, 銀行帳號, 戶名, 電話, 身分證, 備註
        classification = safe_str(row[0])
        member_id      = safe_str(row[1])
        nickname_id    = safe_int(row[2])
        nickname       = safe_str(row[3])
        bank_account   = safe_str(row[4])
        account_holder = safe_str(row[5])
        phone          = safe_str(row[6])
        id_number      = safe_str(row[7])
        notes          = safe_str(row[8])

        if not nickname and not phone:
            continue

        record = {
            'classification': classification or '黑名單',
            'nickname_id':    nickname_id,
            'nickname':       nickname or None,
            'bank_account':   bank_account or None,
            'account_holder': account_holder or None,
            'phone':          phone or None,
            'id_number':      id_number or None,
            'notes':          notes or None,
            'is_blacklist':   True,
        }
        batch.append(record)
        count += 1

        if len(batch) >= 100:
            supabase.table('customers').insert(batch).execute()
            print(f'  ✅ 已匯入 {len(batch)} 筆黑名單')
            batch = []

    if batch:
        supabase.table('customers').insert(batch).execute()
        print(f'  ✅ 已匯入 {len(batch)} 筆黑名單')

    wb.close()
    print(f'  🎉 黑名單匯入完成，共 {count} 筆')


# ─────────────────────────────────────────────
#  匯入訂單（訂單資料.xlsx）
# ─────────────────────────────────────────────

def import_orders():
    print("\n📂 匯入訂單資料...")
    wb = openpyxl.load_workbook(ORDERS_FILE, read_only=True, data_only=True)
    ws = wb.active

    batch = []
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 欄位A-S：日期, 賣紅鑽, 買紅鑽, 暱稱, 轉出, 轉入, 手續費, 結存,
        #          (空), 訂單起, 訂單終, 備註, 日期2, 新編號, 回鑽數量, 銀行帳號, 手續, 時間
        raw_date       = row[0]
        sell_diamonds  = safe_float(row[1])
        buy_diamonds   = safe_float(row[2])
        nickname       = safe_str(row[3])
        transfer_out   = safe_float(row[4])
        transfer_in    = safe_float(row[5])
        fee            = safe_float(row[6])
        balance        = safe_float(row[7])
        order_no_start = safe_str(row[9])
        order_no_end   = safe_str(row[10])
        notes          = safe_str(row[11])
        customer_id    = safe_int(row[13])
        bank_account   = safe_str(row[15])
        time_note      = safe_str(row[17])

        # 跳過沒有暱稱的行（非交易記錄）
        if not nickname:
            continue

        # 跳過初始餘額等系統記錄
        if nickname in ('初始餘額', '寶萌開戶'):
            continue

        # 日期處理
        if isinstance(raw_date, datetime):
            order_date = raw_date.date().isoformat()
        elif isinstance(raw_date, str) and raw_date:
            order_date = raw_date[:10]
        else:
            order_date = None

        record = {
            'date':            order_date,
            'sell_diamonds':   sell_diamonds,
            'buy_diamonds':    buy_diamonds,
            'customer_nickname': nickname,
            'transfer_out':    transfer_out,
            'transfer_in':     transfer_in,
            'fee':             fee,
            'balance':         balance,
            'order_no_start':  order_no_start or None,
            'order_no_end':    order_no_end or None,
            'notes':           notes or None,
            'customer_id_new': customer_id,
            'bank_account':    bank_account or None,
            'time_note':       time_note or None,
        }
        batch.append(record)
        count += 1

        if len(batch) >= 100:
            supabase.table('orders').insert(batch).execute()
            print(f'  ✅ 已匯入 {len(batch)} 筆訂單')
            batch = []

    if batch:
        supabase.table('orders').insert(batch).execute()
        print(f'  ✅ 已匯入 {len(batch)} 筆訂單')

    wb.close()
    print(f'  🎉 訂單匯入完成，共 {count} 筆')


# ─────────────────────────────────────────────
#  主程式
# ─────────────────────────────────────────────

if __name__ == '__main__':
    print('=' * 50)
    print('  寶萌資料庫 Excel 匯入工具')
    print('=' * 50)

    try:
        import_members()
        import_blacklist()
        import_orders()
        print('\n✅ 全部匯入完成！')
    except FileNotFoundError as e:
        print(f'\n❌ 找不到檔案：{e}')
        print('請確認 Excel 檔案放在正確路徑')
    except Exception as e:
        print(f'\n❌ 匯入失敗：{e}')
        raise
